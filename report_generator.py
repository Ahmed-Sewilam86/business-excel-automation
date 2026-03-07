import logging
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def auto_adjust_column_width(worksheet, dataframe):
    """
    Automatically adjust column width based on content.
    """
    for i, column in enumerate(dataframe.columns, 1):
        max_length = max(
            dataframe[column].astype(str).map(len).max(),
            len(column)
        )
        adjusted_width = max_length + 2
        worksheet.column_dimensions[get_column_letter(i)].width = adjusted_width


def format_header(worksheet):
    """
    Apply bold formatting and freeze header row.
    """
    for cell in worksheet[1]:
        cell.font = Font(bold=True)

    worksheet.freeze_panes = "A2"


def format_currency_column(worksheet, dataframe):
    """
    Apply currency formatting to Revenue column if present.
    """
    if "Revenue" in dataframe.columns:
        col_index = dataframe.columns.get_loc("Revenue") + 1
        column_letter = get_column_letter(col_index)

        for row in range(2, len(dataframe) + 2):
            worksheet[f"{column_letter}{row}"].number_format = '"$"#,##0.00'


def add_totals_row(worksheet, dataframe):
    """
    Add totals row for Revenue column if present.
    """
    if "Revenue" in dataframe.columns:
        revenue_index = dataframe.columns.get_loc("Revenue") + 1
        column_letter = get_column_letter(revenue_index)

        total_row = len(dataframe) + 2
        formula = f"=SUM({column_letter}2:{column_letter}{len(dataframe)+1})"

        worksheet[f"{column_letter}{total_row}"] = formula
        worksheet[f"A{total_row}"] = "Total"
        worksheet[f"A{total_row}"].font = Font(bold=True)


def write_sheet(writer, sheet_name, dataframe):
    """
    Write dataframe to Excel and apply formatting.
    """
    dataframe.to_excel(writer, sheet_name=sheet_name, index=False)

    worksheet = writer.book[sheet_name]

    format_header(worksheet)
    auto_adjust_column_width(worksheet, dataframe)
    format_currency_column(worksheet, dataframe)
    add_totals_row(worksheet, dataframe)


def generate_excel_report(output_path: str, processed_data: dict) -> None:
    """
    Generate formatted Excel report.
    """
    try:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:

            write_sheet(writer, "Clean Data", processed_data["clean_data"])
            write_sheet(writer, "Revenue by Product", processed_data["revenue_by_product"])
            write_sheet(writer, "Revenue by Country", processed_data["revenue_by_country"])
            write_sheet(writer, "Monthly Revenue", processed_data["monthly_revenue"])

        logging.info("Excel report generated successfully.")

    except Exception as e:
        logging.error(f"Error generating report: {e}")
        raise